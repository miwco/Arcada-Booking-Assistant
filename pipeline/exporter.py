"""Export the finished plan back into the Bokningsönskemål Excel format.

The original request workbooks are used as TEMPLATES so the format is reproduced
exactly: each workbook is copied, then for every course sheet the header block
(code / name / groups / examiner) is cleaned and the session table is rewritten —
one row per planned session — without changing columns, styling, or the
Groups / Event_type lookup sheets.

The plan comes from the dashboard (resolved placement + your edits). Before
writing anything we validate it; if any conflict is still unresolved (not
approved), we refuse and return the list so the caller can stop and ask.
"""
from __future__ import annotations

import glob
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .dictionaries import APP_DIR, EXPORT_DIR, INFO, load_all
from .normalize import clean_text, normalize_course_code
from .parse_requests import _COLUMN_KEYS, _meta_from_filename
from .scheduler import SLOT_TIMES

EXPORT_ROOT = EXPORT_DIR
SEM_FOLDER = {"autumn 2026": "autumn_2026", "spring 2027": "spring_2027"}
WD_SV = {"Mon": "Måndag", "Tue": "Tisdag", "Wed": "Onsdag", "Thu": "Torsdag", "Fri": "Fredag"}
WD_ORDER = {"Mon": 1, "Tue": 2, "Wed": 3, "Thu": 4, "Fri": 5, "": 9}
SLOT_ORDER = {"AM": 0, "FULL": 1, "PM": 2}
GROUP_RE = re.compile(r"^(Media-\d{2}(-[FLMOP])?|KP-\d{2})$")
# Other-programme group codes (EN-26, Em-26, MSE-26, MTH-26, IT-swe-26, ...) and
# elective labels are valid too — only genuinely malformed codes should warn.
OTHER_GROUP_RE = re.compile(r"^[A-Za-zÅÄÖ][A-Za-z-]*-\d{2}$")
ELECTIVE_GROUPS = {"breddstudier", "öppnayh", "öppna yh", "oppnayh"}


def _group_ok(g):
    return bool(GROUP_RE.match(g)) or g.lower() in ELECTIVE_GROUPS or bool(OTHER_GROUP_RE.match(g))


def _label_rows(ws):
    """Map header-block fields to their row number (value lives in column E)."""
    rows = {}
    for r in range(1, 9):
        a = ws.cell(r, 1).value
        if not isinstance(a, str):
            continue
        low = a.lower()
        if "kurskod" in low:
            rows["code"] = r
        elif "namn / name of study unit" in low or "studieavsnitt" in low:
            rows["name"] = r
        elif "studentgrupper" in low:
            rows["groups"] = r
        elif "examinator" in low:
            rows["examiner"] = r
    return rows


def _table(ws):
    """Return (header_row, {logical_col: col_index})."""
    for r in range(1, 15):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and "vecko" in a.lower() and "nummer" in a.lower():
            colmap = {"week": 1}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                low = v.lower()
                for key, logical in _COLUMN_KEYS:
                    if key in low and logical not in colmap:
                        colmap[logical] = c
            return r, colmap
    return None, {}


def _comment(rec):
    parts = []
    if rec.get("placed_date") and rec.get("slot"):
        parts.append(f"{rec['placed_date']} · kl. {SLOT_TIMES.get(rec['slot'], '')}")
    if rec.get("approvedFlag") and rec.get("kinds"):
        note = rec.get("conflictNote") or ", ".join(rec["kinds"])
        parts.append(f"⚠ APPROVED DOUBLE BOOKING ({note})")
    if rec.get("comments"):
        parts.append(clean_text(rec["comments"]))
    return " | ".join(parts)


def validate(plan):
    """Return (unresolved, warnings). Unresolved = active (un-approved) conflicts."""
    unresolved, bad_groups = [], set()
    for r in plan:
        if r.get("state") == "conflict":
            unresolved.append(f"{r['course']} ({r['cohort']}) — {','.join(r.get('kinds', []))} "
                              f"on {r.get('placed_date')} {r.get('slot')}")
        for g in [x.strip() for x in (r.get("groups") or "").split(";") if x.strip()]:
            if not _group_ok(g):
                bad_groups.add(g)
    warnings = [f"group code '{g}' not in Media-YY-X / KP-YY format (check it)"
                for g in sorted(bad_groups)]
    return unresolved, warnings


# --- self-contained per-cohort export (one course per sheet) ---------------
SEM_FILE = {"autumn 2026": "autumn-2026", "spring 2027": "spring-2027"}
_HEADER = [("Kurskod / Study Unit Code:", "code"),
           ("Studieavsnittets namn / Name of study unit:", "name"),
           ("Studentgrupper / Student groups:", "groups"),
           ("Examinator / Examiner:", "examiner"),
           ("Undervisningsspråk / Tuition language:", "language")]
_TABLE_HDR = ["Veckonummer / Week", "Veckodag / Weekday", "Ggr/vecka", "Minuter / Minutes",
              "Antal studenter", "Önskat klassrum / Room", "Undervisande lärare / Teachers",
              "Bokningstyp / Event type", "Utbildningsprogram", "Kommentar / Comment"]
_HDR_FILL = PatternFill("solid", fgColor="D9D9D9")


def _export_cohort(cohort):
    return bool(re.match(r"Media-\d{2}$", cohort or "")) or (cohort or "").lower() in ("öppnayh", "oppnayh")


def _cohort_file(cohort):
    return "oppnaYH" if cohort.lower() in ("öppnayh", "oppnayh") else cohort.lower()


def _sem_file(sem):
    return SEM_FILE.get(sem, (sem or "").replace(" ", "-").lower())


def _int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 9999


def _sheet_name(name, used):
    base = re.sub(r"[\[\]:*?/\\]", "", clean_text(name) or "Course")[:28].strip() or "Course"
    nm, i = base, 2
    while nm.lower() in used:
        nm = f"{base[:26]} {i}"
        i += 1
    used.add(nm.lower())
    return nm


def _write_course_sheet(ws, code, recs):
    first = recs[0]
    vals = {"code": code, "name": first.get("course", ""), "groups": first.get("groups", ""),
            "examiner": first.get("examiner", ""), "language": first.get("language", "")}
    ws["E1"] = "Fyll i här nedan / Fill in below:"
    ws["E1"].font = Font(italic=True, color="808080")
    for i, (label, key) in enumerate(_HEADER):
        r = i + 2
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 5, vals[key])
    for c, label in enumerate(_TABLE_HDR, 1):
        cell = ws.cell(9, c, label)
        cell.font = Font(bold=True)
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row, approved = 10, 0
    for rec in recs:
        mins = rec.get("minutes")
        time = SLOT_TIMES.get(rec.get("slot", ""), "")
        dur = f"{mins} min" + (f" (kl. {time})" if time else "") if mins not in (None, "") else (f"kl. {time}" if time else "")
        ws.cell(row, 1, rec.get("week"))
        ws.cell(row, 2, WD_SV.get(rec.get("weekday", ""), "").upper())
        ws.cell(row, 3, 1)
        ws.cell(row, 4, dur)
        ws.cell(row, 5, rec.get("num_students", ""))
        ws.cell(row, 6, rec.get("room", ""))
        ws.cell(row, 7, ", ".join(t for t in (rec.get("teachers") or "").split("; ") if t))
        ws.cell(row, 8, rec.get("type") or "Lektion/Lecture")
        ws.cell(row, 9, rec.get("programme") or "Film och media")
        ws.cell(row, 10, _comment(rec))
        if rec.get("approvedFlag") and rec.get("kinds"):
            approved += 1
        row += 1
    for c, w in enumerate([10, 12, 8, 22, 12, 18, 28, 18, 18, 30], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return len(recs), approved


def export_plan(plan, out_dir=None, semester=None, decisions=None):
    """Write the plan into template-based Excel files. `plan` is a list of
    resolved booking records from the dashboard. `semester` (e.g. "autumn 2026"
    or "spring 2027") restricts the export to that batch; None exports both.
    `decisions` (the dashboard's pre-export review list) is saved next to the
    files as a human-readable decision log so every change is on record."""
    out_dir = out_dir or EXPORT_ROOT
    if semester:
        plan = [r for r in plan if r.get("semester") == semester]
    unresolved, warnings = validate(plan)
    if unresolved:
        return {"ok": False, "error": "unresolved_conflicts",
                "unresolved": unresolved,
                "message": f"{len(unresolved)} unresolved conflict(s) — resolve or approve them, "
                           f"then export again."}

    # group the plan -> {(cohort, semester): {course_code: [recs]}}, Film & Media only
    by_cohort = {}
    for r in plan:
        if r.get("external") or r.get("context") or not r.get("placed_date"):
            continue
        if not _export_cohort(r.get("cohort", "")):       # Media-YY + ÖppnaYH only
            continue
        ck = (r["cohort"], r["semester"])
        by_cohort.setdefault(ck, {}).setdefault(r.get("course_code") or r.get("course") or "?", []).append(r)

    os.makedirs(out_dir, exist_ok=True)
    files_out, total_rows, approved = [], 0, 0
    for (cohort, sem), courses in sorted(by_cohort.items()):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sheet_rows, used = 0, set()
        for code, recs in sorted(courses.items()):
            recs.sort(key=lambda r: (_int(r.get("week")), WD_ORDER.get(r.get("weekday", ""), 9),
                                     SLOT_ORDER.get(r.get("slot", ""), 9)))
            ws = wb.create_sheet(_sheet_name(recs[0].get("course") or code, used))
            n, na = _write_course_sheet(ws, code, recs)
            sheet_rows += n
            approved += na
        out_path = os.path.join(out_dir, f"{_cohort_file(cohort)}-{_sem_file(sem)}.xlsx")
        wb.save(out_path)
        wb.close()
        files_out.append({"file": os.path.relpath(out_path, APP_DIR), "rows": sheet_rows})
        total_rows += sheet_rows

    log_rel = _write_decision_log(out_dir, semester, decisions, files_out, total_rows, approved)

    return {"ok": True, "files": files_out, "total_rows": total_rows,
            "approved_double_bookings": approved, "warnings": warnings,
            "semester": semester or "both semesters",
            "decision_log": log_rel,
            "export_dir": os.path.relpath(out_dir, APP_DIR)}


def _write_decision_log(out_dir, semester, decisions, files_out, total_rows, approved):
    """Save a JSON + plain-text record of every change made before this export."""
    import datetime
    import json
    decisions = decisions or []
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = SEM_FOLDER.get(semester, "both") if semester else "both"
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.join(out_dir, f"decision_log_{tag}_{stamp}")
    meta = {"exported_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "semester": semester or "both semesters", "sessions_written": total_rows,
            "approved_double_bookings": approved,
            "files": [f["file"] for f in files_out], "decisions": decisions}
    with open(base + ".json", "w", encoding="utf-8") as fh:
        json.dump(meta, fh, ensure_ascii=False, indent=2)
    lines = [f"Booking Assistant — decision log",
             f"Exported: {meta['exported_at']}",
             f"Batch: {meta['semester']}  ·  {total_rows} sessions  ·  {approved} approved double-booking(s)",
             f"Files: {', '.join(meta['files']) or '(none)'}", "",
             f"Changes made before export ({len(decisions)}):"]
    if decisions:
        for dcn in decisions:
            lines.append(f"  - [{dcn.get('type','change')}] {dcn.get('text','')}")
    else:
        lines.append("  (no manual changes — exported as imported)")
    with open(base + ".txt", "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(lines) + "\n")
    return os.path.relpath(base + ".txt", APP_DIR)
