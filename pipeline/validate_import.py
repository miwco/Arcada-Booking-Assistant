"""Import-validation for the ORIGINAL Bokningsönskemål files.

The originals live in _info/original_bokningsönskemål_spring_2027/ — one workbook
per examiner (filename = examiner). They contain many blank template tabs and
messy / contradictory data. This module:

  * skips empty/template tabs (only real courses are processed),
  * identifies the examiner from the filename (a clue to which courses belong),
  * reads the structured fields AND the 'Other comments' field, and
  * runs validators that flag issues and, where possible, SUGGEST a correction:
      - structured fields contradicting the comment (e.g. Gatufotografering:
        5×45 min but the comment says "Wednesdays 10:00–15:00" → 1×300 min),
      - unknown / messy course codes, course-code↔name mismatches,
      - bad / missing group codes,
      - examiner field disagreeing with the file owner or the usual examiner,
      - messy text and unusual values.

Output: findings as a structure (for the interactive report) + a flat CSV.
Nothing is changed automatically — corrections are suggestions to approve.
"""
from __future__ import annotations

import csv
import glob
import json
import os
import re
import unicodedata

import openpyxl

from .dictionaries import COHORT_YEARS, INFO, ROOT, load_all
from .normalize import (clean_text, is_probably_name, normalize_course_code, normalize_group_code,
                        parse_comment, parse_groups, split_people)

GROUP_FORM = re.compile(r"^[A-Za-zÅÄÖ][A-Za-zÅÄÖ-]*-\d{2}(-[A-Za-z]{1,2})?$")  # any PROGRAMME-YY
from .parse_requests import _COLUMN_KEYS

ORIG_DIR = os.path.join(INFO, "original_bokningsönskemål_spring_2027")


def _xlsx(pattern):
    """Glob .xlsx files, skipping Excel lock/temp files (~$...)."""
    return [f for f in glob.glob(pattern, recursive=True)
            if not os.path.basename(f).startswith("~$")]
OUT = os.path.join(ROOT, "output")
SKIP_SHEETS = {"Groups", "Event_type", "Example"}
TIME_RANGE = re.compile(r"(\d{1,2})[.:](\d{2})\s*[-–till\s]+\s*(\d{1,2})[.:](\d{2})")
TIME_RANGE_H = re.compile(r"\bkl\.?\s*(\d{1,2})\s*[-–]\s*(\d{1,2})\b")
GROUP_OK = re.compile(r"^(Media-\d{2}(-[FLMOP])?|KP-\d{2})$")
# Elective / öppnaYH / breddstudier courses have no specific student group.
ELECTIVE_CODES = {"FM-3-007", "MK-2-155", "YH-4-032"}  # Gatufotografering, Game design, Cilect
ELECTIVE_RE = re.compile(r"bredd|öppna|oppna|öppen|öppnayh", re.I)
ELECTIVE_GROUPS = {"breddstudier", "öppnayh", "öppna yh", "oppnayh"}


def _strip(s):
    return "".join(c for c in unicodedata.normalize("NFD", s or "") if unicodedata.category(c) != "Mn").lower()


def examiner_from_filename(fname, d):
    """Map a filename to a canonical examiner using names + nickname aliases."""
    f = _strip(fname)
    best = None
    for alias, canon in d.teacher_alias.items():
        if _strip(alias) and _strip(alias) in f:
            if best is None or len(alias) > best[0]:
                best = (len(alias), canon)
    for name in d.teachers:                       # also try first/last name tokens
        for tok in name.split():
            if len(tok) >= 3 and _strip(tok) in f and (best is None or len(tok) > best[0]):
                best = (len(tok), name)
    return best[1] if best else None


def comment_duration(text):
    """Minutes implied by a time range in free text, or None."""
    t = str(text or "")
    m = TIME_RANGE.search(t)
    if m:
        sh, sm, eh, em = map(int, m.groups())
        d = (eh * 60 + em) - (sh * 60 + sm)
        return d if 0 < d <= 720 else None
    m = TIME_RANGE_H.search(t)
    if m:
        sh, eh = int(m.group(1)), int(m.group(2))
        if 6 <= sh < eh <= 22:
            return (eh - sh) * 60
    return None


def _label_rows(rows):
    out = {}
    for r in range(min(9, len(rows))):
        a = rows[r][0] if rows[r] else None
        if not isinstance(a, str):
            continue
        low = a.lower()
        if "kurskod" in low: out["code"] = r
        elif "namn / name of study unit" in low or "studieavsnitt" in low: out["name"] = r
        elif "studentgrupper" in low: out["groups"] = r
        elif "examinator" in low: out["examiner"] = r
    return out


def _find_table(rows):
    for i, r in enumerate(rows):
        if r and isinstance(r[0], str) and "vecko" in r[0].lower() and "nummer" in r[0].lower():
            cm = {"week": 0}
            for j, cell in enumerate(r):
                if isinstance(cell, str):
                    low = cell.lower()
                    for key, logical in _COLUMN_KEYS:
                        if key in low and logical not in cm:
                            cm[logical] = j
            return i, cm
    return None, {}


def _val(rows, r, col=4):
    return clean_text(rows[r][col]) if r is not None and r < len(rows) and len(rows[r]) > col else ""


def validate_course(rows, sheet, fname, owner, d, ref):
    lr = _label_rows(rows)
    code_raw = _val(rows, lr.get("code"))
    name_raw = _val(rows, lr.get("name"))
    hdr, cm = _find_table(rows)
    data = []
    if hdr is not None:
        for ri in range(hdr + 1, len(rows)):
            row = rows[ri]
            if not row:
                continue
            get = lambda lo: clean_text(row[cm[lo]]) if lo in cm and cm[lo] < len(row) else ""
            wk = get("week")
            if not wk and not get("teachers") and not get("comments"):
                continue
            data.append({"week": wk, "weekday": get("weekday"), "tpw": get("times_per_week"),
                         "minutes": get("minutes"), "room": get("room"), "teachers": get("teachers"),
                         "type": get("booking_type"), "content": get("content"), "comments": get("comments")})
    # ---- is this a real course tab? ------------------------------------
    if not code_raw and not name_raw and not data:
        return None

    findings = []

    def add(field, sev, issue, current="", suggested="", conf="med"):
        findings.append({"field": field, "severity": sev, "issue": issue,
                         "current": str(current), "suggested": str(suggested), "confidence": conf})

    # ---- course code ----------------------------------------------------
    code, cnotes = normalize_course_code(code_raw)
    canon, cname, lnote = d.lookup_course(code) if code else (None, None, "")
    for n in cnotes:
        add("course_code", "info", n, code_raw, code)
    if code and not canon:
        # only worth flagging when there's an established catalog (then it may be a typo);
        # on a fresh import the catalog is generated FROM these codes, so don't repeat it
        # for every course — the import summary reports "N courses added" once.
        if len(d.courses) > 5:
            add("course_code", "warn",
                f"course code not in your catalog: {code} (typo? otherwise it is added)", code_raw, "", "low")
    elif canon and canon != code:
        add("course_code", "info", "code normalized to dictionary form", code_raw, canon, "high")
    if canon and cname and name_raw and cname.lower() != name_raw.lower():
        add("course_name", "info", "course name differs from the dictionary", name_raw, cname, "med")

    # ---- groups ---------------------------------------------------------
    groups_raw = _val(rows, lr.get("groups"))
    elective = (canon in ELECTIVE_CODES) or bool(ELECTIVE_RE.search(groups_raw + " " + name_raw)) or \
        any(ELECTIVE_RE.search((r.get("content", "") + " " + r.get("comments", ""))) for r in data)
    if not groups_raw:
        if elective:
            add("groups", "info", "elective (öppnaYH / breddstudier) — no specific student group needed",
                "", "Breddstudier", "low")
        else:
            add("groups", "warn", "no student groups given", "", "", "low")
    else:
        gs = parse_groups(groups_raw)[0]
        # suggest cleaned/normalised codes (Media-2025 -> Media-25, Media-2 -> a cohort, …)
        norm, notes = [], []
        for g in gs:
            nc, note = normalize_group_code(g, COHORT_YEARS)
            norm.append(nc or g)
            if note:
                notes.append(note)
        norm_str = "; ".join(dict.fromkeys(n for n in norm if n))
        if norm_str and norm_str != "; ".join(gs):
            add("groups", "warn", "group codes look inconsistent — normalise to " + norm_str,
                groups_raw, norm_str, "medium")
        for g in norm:
            if g and not GROUP_FORM.match(g) and g.lower() not in ELECTIVE_GROUPS:
                add("groups", "warn", f"group code not in PROGRAMME-YY form: {g}", groups_raw, "", "low")

    # ---- examiner (the in-sheet 'Examinator' field is authoritative) ----
    exam_sheet = _val(rows, lr.get("examiner"))
    exam_canon, en = d.lookup_teacher(exam_sheet) if exam_sheet else ("", "")
    ref_exam = ref.get(canon) if canon else None
    if not exam_sheet:
        hint = (", ".join(sorted(ref_exam)) if ref_exam else owner) or ""
        add("examiner", "warn", "Examinator field is empty (it is the source of truth for the examiner)",
            "", hint, "low")
    elif exam_canon and exam_canon != exam_sheet:
        add("examiner", "info", "examiner name normalized to the team list", exam_sheet, exam_canon, "med")

    # ---- teacher cells: flag notes/instructions written as a "teacher" ----
    seen_notes = set()
    for row in data:
        for t in split_people(row["teachers"]):
            ct = clean_text(t)
            if ct and not is_probably_name(ct) and ct.lower() not in seen_notes:
                seen_notes.add(ct.lower())
                add("teachers", "warn",
                    f"the teacher cell holds a note, not a name: '{ct}' — left off the teacher list; "
                    f"tell the app what it means (e.g. co-scheduled with another group)", ct, "", "med")

    # ---- fields vs 'Other comments' (the marquee check) -----------------
    fix = None
    for row in data:
        dur = comment_duration(row["comments"])
        try:
            mins = int(float(row["minutes"])) if row["minutes"] else None
        except ValueError:
            mins = None
        try:
            tpw = int(float(row["tpw"])) if row["tpw"] else None
        except ValueError:
            tpw = None
        if dur and ((mins and abs(mins - dur) > 5) or (tpw and tpw > 1 and dur >= 120)):
            fix = {"dur": dur, "mins": mins, "tpw": tpw, "comment": row["comments"]}
            break
    if fix:
        add("minutes/frequency", "warn",
            f"structured fields contradict the comment '{fix['comment']}': "
            f"{fix['tpw']}×{fix['mins']} min, but the comment means one session of {fix['dur']} min",
            f"{fix['tpw']}×{fix['mins']} min", f"1×{fix['dur']} min", "high")

    # ---- unusual values -------------------------------------------------
    for row in data:
        try:
            mins = int(float(row["minutes"])) if row["minutes"] else None
            tpw = int(float(row["tpw"])) if row["tpw"] else None
        except ValueError:
            mins = tpw = None
        if mins and mins < 60 and tpw and tpw >= 4:
            add("minutes/frequency", "warn",
                f"unusual: {tpw} times/week of only {mins} min — check this is intended",
                f"{tpw}×{mins} min", "", "low")
            break

    # ---- messy text -----------------------------------------------------
    if any(ch in (code_raw + name_raw) for ch in "\t\n"):
        add("text", "info", "tabs/newlines in code or name", code_raw + " / " + name_raw, "", "low")

    return {"file": os.path.basename(fname), "owner": owner or "?", "sheet": sheet,
            "code": canon or code, "code_raw": code_raw, "name": cname or name_raw,
            "examiner": exam_canon or exam_sheet, "groups": groups_raw, "n_rows": len(data),
            "findings": findings, "rows": data}


def build_examiner_ref(d):
    """course_code -> set(examiner) learned from the cleaned reference files."""
    ref = {}
    for f in _xlsx(os.path.join(INFO, "bokningsönskemålen_2026_2027", "**", "*.xlsx")):
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        for ws in wb.worksheets:
            if ws.title in SKIP_SHEETS:
                continue
            rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
            lr = _label_rows(rows)
            code = _val(rows, lr.get("code"))
            code, _ = normalize_course_code(code)
            canon, _n, _l = d.lookup_course(code) if code else (None, None, "")
            exam = _val(rows, lr.get("examiner"))
            ec, _ = d.lookup_teacher(exam) if exam else ("", "")
            if canon and ec:
                ref.setdefault(canon, set()).add(ec)
        wb.close()
    return ref


def validate_all(folder=None):
    folder = folder or ORIG_DIR
    d = load_all()
    ref = build_examiner_ref(d)
    files = []
    for f in sorted(_xlsx(os.path.join(folder, "*.xlsx"))):
        owner = examiner_from_filename(os.path.basename(f), d)
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        courses = []
        for ws in wb.worksheets:
            if ws.title in SKIP_SHEETS:
                continue
            res = validate_course(list(ws.iter_rows(values_only=True)), ws.title, f, owner, d, ref)
            if res:
                courses.append(res)
        wb.close()
        files.append({"file": os.path.basename(f), "owner": owner or "(unknown)",
                      "courses": courses})
    return files


def write_csv(files):
    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, "import_validation.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["file", "examiner", "course_sheet", "course_code", "field",
                    "severity", "issue", "current", "suggested", "confidence"])
        for fl in files:
            for c in fl["courses"]:
                for fd in c["findings"]:
                    w.writerow([fl["file"], fl["owner"], c["sheet"], c["code"], fd["field"],
                                fd["severity"], fd["issue"], fd["current"], fd["suggested"], fd["confidence"]])


REPORT = r"""<!DOCTYPE html><html lang="sv"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Import validation · Bokningsönskemål</title>
<style>
 :root{--bg:#0f1419;--panel:#1a212b;--panel2:#222c38;--line:#33414f;--txt:#e6edf3;--muted:#8b9aa8;
  --accent:#4da3ff;--err:#fb7185;--warn:#facc15;--info:#64748b;--ok:#34d399}
 *{box-sizing:border-box}body{margin:0;font:14px/1.45 "Segoe UI",system-ui,sans-serif;background:var(--bg);color:var(--txt)}
 header{position:sticky;top:0;background:var(--panel);border-bottom:1px solid var(--line);padding:12px 20px;
  display:flex;gap:14px;align-items:center;flex-wrap:wrap;z-index:5}
 header h1{font-size:17px;margin:0}
 main{padding:16px 20px;max-width:1000px}
 button{background:var(--panel2);color:var(--txt);border:1px solid var(--line);border-radius:7px;padding:7px 12px;cursor:pointer;font:inherit}
 button.primary{background:var(--accent);color:#04121f;border-color:var(--accent);font-weight:600}
 .file{margin:18px 0 6px;font-size:15px;color:var(--accent);font-weight:600}
 .file span{color:var(--muted);font-weight:400;font-size:13px}
 .course{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:12px 14px;margin:8px 0}
 .course h3{margin:0 0 2px;font-size:14px}
 .meta{color:var(--muted);font-size:12.5px;margin-bottom:8px}
 .pill{display:inline-block;background:var(--panel2);border:1px solid var(--line);border-radius:20px;padding:1px 9px;font-size:11px;margin-left:6px;color:var(--muted)}
 .find{border-left:3px solid var(--info);background:var(--panel2);border-radius:7px;padding:8px 10px;margin:6px 0}
 .find.error{border-color:var(--err)}.find.warn{border-color:var(--warn)}.find.info{border-color:var(--info)}
 .sev{font-size:10.5px;font-weight:800;text-transform:uppercase;letter-spacing:.04em}
 .sev.error{color:var(--err)}.sev.warn{color:var(--warn)}.sev.info{color:var(--muted)}
 .sug{margin-top:5px;font-size:13px}
 .cur{color:var(--err)}.arrow{color:var(--muted)}.new{color:var(--ok);font-weight:600}
 .act{margin-top:6px;display:flex;gap:6px}
 .act .b{padding:3px 10px;font-size:12px;border-radius:14px}
 .b.on{background:var(--ok);color:#04121f;border-color:var(--ok);font-weight:700}
 .b.off{background:var(--err);color:#1a0a0d;border-color:var(--err);font-weight:700}
 .clean{color:var(--ok);font-size:12.5px}
 .muted{color:var(--muted)} .count{font-variant-numeric:tabular-nums}
</style></head><body>
<header><h1>🔎 Import validation <span class="muted" style="font-weight:400">· original Bokningsönskemål · spring 2027</span></h1>
 <div id="summary" class="muted"></div>
 <div style="margin-left:auto"><button class="primary" id="dl">⬇ Download approved corrections</button></div>
</header><main id="main"></main>
<script>
const FILES=__FILES__;
const $=s=>document.querySelector(s);
const esc=s=>(s==null?"":String(s)).replace(/[&<>"]/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));
const KEY="ba_import_approvals";
let AP=JSON.parse(localStorage.getItem(KEY)||"{}");
const fid=(fi,ci,k)=>fi+"|"+ci+"|"+k;
function setAp(id,v){AP[id]=v;localStorage.setItem(KEY,JSON.stringify(AP));render();}
function render(){
  let total=0,withFix=0,approved=0;
  let html="";
  FILES.forEach((f,fi)=>{
    html+=`<div class="file">📄 File by ${esc(f.owner)} <span>— ${esc(f.file)} · ${f.courses.length} course(s) · examiner is taken from each course's Examinator field</span></div>`;
    f.courses.forEach((c,ci)=>{
      total+=c.findings.length;
      html+=`<div class="course"><h3>${esc(c.name||c.sheet)} <span class="pill">${esc(c.code||"no code")}</span>`+
        `<span class="pill">${c.n_rows} rows</span></h3>`+
        `<div class="meta">examiner field: ${esc(c.examiner||"—")} · groups: ${esc(c.groups||"—")}</div>`;
      if(!c.findings.length){ html+=`<div class="clean">✓ no issues found</div>`; }
      c.findings.forEach((fd,k)=>{
        const id=fid(fi,ci,k); const hasFix=!!fd.suggested; if(hasFix)withFix++;
        const a=AP[id]; if(a==="yes")approved++;
        html+=`<div class="find ${fd.severity}"><span class="sev ${fd.severity}">${fd.severity}</span> `+
          `<b>${esc(fd.field)}</b> · <span class="muted">${esc(fd.confidence)} confidence</span><br>${esc(fd.issue)}`;
        if(hasFix) html+=`<div class="sug"><span class="cur">${esc(fd.current||"—")}</span> <span class="arrow">→</span> <span class="new">${esc(fd.suggested)}</span></div>`+
          `<div class="act"><button class="b ${a==="yes"?"on":""}" onclick="setAp('${id}','yes')">${a==="yes"?"✓ Approved":"Approve fix"}</button>`+
          `<button class="b ${a==="no"?"off":""}" onclick="setAp('${id}','no')">${a==="no"?"✗ Rejected":"Reject"}</button></div>`;
        html+=`</div>`;
      });
      html+=`</div>`;
    });
  });
  $("#main").innerHTML=html;
  $("#summary").innerHTML=`<span class="count">${total}</span> findings · <span class="count">${withFix}</span> with a suggested fix · <span class="count" style="color:var(--ok)">${approved}</span> approved`;
}
$("#dl").onclick=()=>{
  const out=[];
  FILES.forEach((f,fi)=>f.courses.forEach((c,ci)=>c.findings.forEach((fd,k)=>{
    if(AP[fid(fi,ci,k)]==="yes"&&fd.suggested) out.push({file:f.file,examiner:f.owner,sheet:c.sheet,code:c.code,field:fd.field,from:fd.current,to:fd.suggested,issue:fd.issue});})));
  const blob=new Blob([JSON.stringify({approved:out},null,2)],{type:"application/json"});
  const a=document.createElement("a");a.href=URL.createObjectURL(blob);a.download="approved_corrections.json";a.click();
};
render();
</script></body></html>"""


def write_report(files):
    os.makedirs(OUT, exist_ok=True)
    html = REPORT.replace("__FILES__", json.dumps(files, ensure_ascii=False))
    with open(os.path.join(OUT, "import_validation.html"), "w", encoding="utf-8") as fh:
        fh.write(html)


if __name__ == "__main__":
    files = validate_all()
    write_csv(files)
    write_report(files)
    nc = sum(len(f["courses"]) for f in files)
    nf = sum(len(c["findings"]) for f in files for c in f["courses"])
    print(f"{len(files)} files, {nc} real courses, {nf} findings")
    for f in files:
        fc = sum(len(c["findings"]) for c in f["courses"])
        print(f"  {f['file']}  (file by: {f['owner']})  — {len(f['courses'])} courses, {fc} findings")
