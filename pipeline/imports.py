"""Generous Excel import for teachers, courses and groups + downloadable templates.

"Generous" = the column headers don't have to match the template exactly. Headers
are normalised and matched against a list of synonyms (English/Swedish/Finnish), so
"Teacher", "Lärare", "Name", "Namn", "Lecturer" all map to the name column, etc.
Everything is parsed and validated and returned for the user to review BEFORE it is
saved — nothing is written by these functions.
"""
from __future__ import annotations

import io
import os
import re

import openpyxl

# ---- synonyms (normalised: lowercased, non-alphanumerics stripped) ---------
TEACHER_SYN = {
    "name": ["name", "teacher", "teachername", "lärare", "larare", "namn", "fullname",
             "lecturer", "person", "opettaja", "nimi"],
    "aliases": ["alias", "aliases", "nickname", "smeknamn", "alternative", "alternativ",
                "alternativa", "altnames", "alternativespelling", "spelling", "stavning",
                "othernames", "kallas", "lempinimi"],
}
COURSE_SYN = {
    "code": ["code", "kod", "kurskod", "coursecode", "courseid", "id", "kurskoder"],
    "name": ["name", "namn", "course", "kurs", "coursename", "kursnamn", "title", "titel",
             "studieavsnitt", "studyunit", "nameofstudyunit"],
    "ects": ["ects", "sp", "credits", "op", "studiepoäng", "studiepoang", "hp", "creditpoints"],
    "aliases": ["alias", "aliases", "alternative", "alternativ", "altname", "othernames",
                "alternativnamn"],
    "program": ["program", "programme", "group", "grupp", "programkod", "programmecode",
                "programgrupp", "utbildning"],
}
GROUP_SYN = {
    "code": ["group", "grupp", "groupcode", "gruppkod", "code", "kod", "studentgrupp", "klass"],
    "program": ["program", "programme", "programkod", "programmecode", "utbildning"],
    "name": ["name", "namn", "programname", "description", "beskrivning", "programnamn"],
    "year": ["year", "år", "ar", "yy", "intake", "cohort", "årskurs", "arskurs", "vuosi"],
}


def _norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _match_columns(headers, syn):
    """Map logical field -> column index. Each column is used by at most one field.
    Pass 1 = exact header match (most reliable); pass 2 = a synonym appears INSIDE the
    header (e.g. 'teachername' contains 'name'). We deliberately don't match a header
    that is merely a substring of a synonym ('namn' inside 'smeknamn'), which would
    wrongly tie the name column to the alias field."""
    norm = [_norm(h) for h in headers]
    out, used = {}, set()
    for field, words in syn.items():
        wn = [_norm(w) for w in words]
        idx = next((i for i, h in enumerate(norm) if i not in used and h and h in wn), None)
        if idx is not None:
            out[field] = idx
            used.add(idx)
    for field, words in syn.items():
        if field in out:
            continue
        wn = [_norm(w) for w in words]
        idx = next((i for i, h in enumerate(norm)
                    if i not in used and h and any(w in h for w in wn)), None)
        if idx is not None:
            out[field] = idx
            used.add(idx)
    return out


def _rows(path_or_bytes, syn):
    """Open the first non-empty sheet, find the header row (best synonym match in the
    first 10 rows), and return (colmap, [row-as-list], header_labels)."""
    src = io.BytesIO(path_or_bytes) if isinstance(path_or_bytes, (bytes, bytearray)) else path_or_bytes
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = next((w for w in wb.worksheets if w.max_row and w.max_column), wb.active)
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    rows = [r for r in rows if any(v not in (None, "") for v in r)]
    if not rows:
        return {}, [], []
    best_i, best_map = 0, {}
    for i, r in enumerate(rows[:10]):
        m = _match_columns(r, syn)
        if len(m) > len(best_map):
            best_map, best_i = m, i
    headers = rows[best_i]
    return best_map, rows[best_i + 1:], headers


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


# ---- teachers --------------------------------------------------------------
def parse_teachers(path_or_bytes):
    cm, rows, headers = _rows(path_or_bytes, TEACHER_SYN)
    out, seen = [], set()
    for r in rows:
        name = _cell(r, cm.get("name")) or (_cell(r, 0) if "name" not in cm else "")
        if not name:
            continue
        aliases = [a.strip() for a in re.split(r"[;,/]", _cell(r, cm.get("aliases"))) if a.strip()]
        issue = ""
        if name.lower() in seen:
            issue = "duplicate name"
        seen.add(name.lower())
        out.append({"name": name, "aliases": aliases, "issue": issue})
    return {"matched_columns": {k: headers[v] for k, v in cm.items() if v < len(headers)},
            "items": out, "count": len(out)}


# ---- courses ---------------------------------------------------------------
def parse_courses(path_or_bytes):
    cm, rows, headers = _rows(path_or_bytes, COURSE_SYN)
    out, seen = [], set()
    for r in rows:
        code = _cell(r, cm.get("code")) or (_cell(r, 0) if "code" not in cm else "")
        code = code.upper().strip()
        if not code:
            continue
        ects = _cell(r, cm.get("ects"))
        issues = []
        if code in seen:
            issues.append("duplicate code")
        if ects and not re.match(r"^\d+([.,]\d+)?$", ects):
            issues.append("ECTS not a number")
        seen.add(code)
        out.append({"code": code, "name": _cell(r, cm.get("name")),
                    "ects": ects.replace(",", "."), "aliases": _cell(r, cm.get("aliases")),
                    "program": _cell(r, cm.get("program")), "notes": "",
                    "issue": "; ".join(issues)})
    return {"matched_columns": {k: headers[v] for k, v in cm.items() if v < len(headers)},
            "items": out, "count": len(out)}


# ---- groups ----------------------------------------------------------------
_GROUP_RE = re.compile(r"^[A-Za-zÅÄÖåäö]+-\d{2}(-[A-Za-z]{1,2})?$")


def parse_groups(path_or_bytes):
    cm, rows, headers = _rows(path_or_bytes, GROUP_SYN)
    groups, progs = [], []
    for r in rows:
        code = _cell(r, cm.get("code"))
        prog = _cell(r, cm.get("program"))
        name = _cell(r, cm.get("name"))
        year = _cell(r, cm.get("year"))
        if code:
            # build a full code if only a program + year were given
            if not re.search(r"-\d", code) and year:
                yy = re.sub(r"\D", "", year)[-2:]
                if yy:
                    code = f"{code}-{yy}"
            issue = "" if _GROUP_RE.match(code) else "unusual format (expected e.g. FT-26)"
            groups.append({"code": code, "issue": issue})
        if prog:
            progs.append({"code": prog.strip(), "name": name})
    return {"matched_columns": {k: headers[v] for k, v in cm.items() if v < len(headers)},
            "groups": groups, "programs": progs, "count": len(groups) + len(progs)}


# ---- downloadable templates ------------------------------------------------
_TEMPLATES = {
    "teachers": (["Name", "Aliases (optional, separate with ;)"],
                 [["Example Teacher", "ET; E. Teacher"], ["Second Example", ""]]),
    "courses": (["Code", "Name", "ECTS", "Aliases (optional)", "Program (optional)"],
                [["MK-2-131", "Produktion i broadcastmiljö", "5", "", "Media"],
                 ["FT-1-001", "Introduktion", "5", "", "FT"]]),
    "groups": (["Program code", "Program name", "Year (YY)", "Group code (optional)"],
               [["FT", "Fysioterapi", "26", "FT-26"], ["Media", "Film och media", "26", "Media-26"]]),
}


def _template_wb(kind):
    headers, examples = _TEMPLATES[kind]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = kind.capitalize()
    ws.append(headers)
    for ex in examples:
        ws.append(ex)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, len(h) + 2)
    return wb


def template_bytes(kind):
    buf = io.BytesIO()
    _template_wb(kind).save(buf)
    return buf.getvalue()


def write_templates(folder):
    os.makedirs(folder, exist_ok=True)
    names = {"teachers": "teacher_template.xlsx", "courses": "course_template.xlsx",
             "groups": "group_template.xlsx"}
    for kind, fn in names.items():
        _template_wb(kind).save(os.path.join(folder, fn))
    return list(names.values())
