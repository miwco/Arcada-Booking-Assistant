"""Apply approved import corrections and feed the planner from the originals.

Flow:
  1. Take the approved corrections (from the validation report's "Download
     approved corrections", default output/approved_corrections.json — or pass
     --all to apply every suggested fix).
  2. Apply them to copies of the ORIGINAL per-examiner files (fix minutes/
     frequency, names, codes, groups, examiner) -> export/validated_spring_2027/
     corrected_originals/.
  3. Parse those corrected files with the normal pipeline (reusing all the
     cleaning/normalisation), regroup each course to its cohort (from its
     groups; electives -> ÖppnaYH), and combine with the cleaned AUTUMN data.
  4. Write output/bookings_2026_2027.csv (the planner's source) and regenerate
     output/dashboard.html.

The corrected planner data then drives the dashboard; the final booker Excel is
produced from there with the existing Export Excel button.
"""
from __future__ import annotations

import csv
import glob
import json
import os
import re
import sys
from dataclasses import asdict, fields

import openpyxl

from .dictionaries import COHORT_YEARS, EXPORT_DIR, IMPORT_DIR, INFO, OUTPUT_DIR, load_all
from .exporter import _label_rows, _table
from .normalize import normalize_group_code
from .parse_requests import Booking, parse_all, parse_file
from .validate_import import ELECTIVE_CODES, ELECTIVE_RE, ORIG_DIR, _xlsx, validate_all

_PROG_YEAR = re.compile(r"^([A-Za-zÅÄÖ][A-Za-zÅÄÖ-]*?-\d{2})")

OUT = OUTPUT_DIR
# the import writes corrected copies of the uploaded files as a working step — keep
# these out of export/ (which should hold ONLY the final booker files)
VAL_DIR = os.path.join(IMPORT_DIR, "_corrected")
CORR_DIR = VAL_DIR
SPRING_SEM, SPRING_AY = "spring 2027", "2026-2027"


def _to_freq_minutes(s):
    m = re.search(r"(\d+)\s*[×x]\s*(\d+)", s or "")
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def load_corrections(arg):
    """Return {filebasename: {sheet: [{field,to}]}}."""
    if arg == "--all":
        out = {}
        for f in validate_all():
            for c in f["courses"]:
                for fd in c["findings"]:
                    if fd["suggested"]:
                        out.setdefault(f["file"], {}).setdefault(c["sheet"], []).append(
                            {"field": fd["field"], "to": fd["suggested"]})
        return out
    path = arg or os.path.join(OUT, "approved_corrections.json")
    with open(path, encoding="utf-8-sig") as fh:
        data = json.load(fh)
    out = {}
    for c in data.get("approved", []):
        out.setdefault(c["file"], {}).setdefault(c["sheet"], []).append(
            {"field": c["field"], "to": c["to"]})
    return out


def apply_to_file(path, per_sheet):
    """Write a corrected copy of one original file; return its path."""
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        cs = per_sheet.get(ws.title)
        if not cs:
            continue
        lr = _label_rows(ws)
        hdr, cm = _table(ws)
        for c in cs:
            f, to = c["field"], c["to"]
            if f == "course_name" and "name" in lr:
                ws.cell(lr["name"], 5).value = to
            elif f == "course_code" and "code" in lr:
                ws.cell(lr["code"], 5).value = to
            elif f == "groups" and "groups" in lr:
                ws.cell(lr["groups"], 5).value = to
            elif f == "examiner" and "examiner" in lr:
                ws.cell(lr["examiner"], 5).value = to
            elif f == "minutes/frequency" and hdr is not None:
                n, mins = _to_freq_minutes(to)
                if mins:
                    for r in range(hdr + 1, ws.max_row + 1):
                        if ws.cell(r, cm.get("week", 1)).value in (None, ""):
                            continue
                        if "times_per_week" in cm:
                            ws.cell(r, cm["times_per_week"]).value = n or 1
                        if "minutes" in cm:
                            ws.cell(r, cm["minutes"]).value = mins
    os.makedirs(CORR_DIR, exist_ok=True)
    outp = os.path.join(CORR_DIR, os.path.basename(path))
    wb.save(outp)
    wb.close()
    return outp


def cohort_of(b):
    """The cohort a booking belongs to — any programme, not just Media. Derived by
    normalising the booking's group codes to PROG-YY (Media keeps the whole-year
    cohort, dropping the -track). Falls back to the raw cell, then electives."""
    for source in (b.groups or "", b.groups_raw or ""):
        for g in re.split(r"[;,]", source):
            norm, _ = normalize_group_code(g, COHORT_YEARS)
            m = _PROG_YEAR.match(norm)
            if m:
                return m.group(1)
    if b.course_code in ELECTIVE_CODES or ELECTIVE_RE.search(
            (b.groups_raw or "") + " " + (b.course_name or "") + " " + (b.groups or "")):
        return "ÖppnaYH"
    return None  # no usable group code at all


def _semester_of(week):
    try:
        return "autumn 2026" if int(week) >= 30 else SPRING_SEM
    except (ValueError, TypeError):
        return SPRING_SEM


def corrections_from_list(approved):
    """[{file,sheet,field,to}, ...] -> {file: {sheet: [{field,to}]}}."""
    out = {}
    for c in (approved or []):
        out.setdefault(c["file"], {}).setdefault(c["sheet"], []).append(
            {"field": c["field"], "to": c["to"]})
    return out


def _suggest_group(raw):
    for g in re.split(r"[;,]", raw or ""):
        nc, _ = normalize_group_code(g, COHORT_YEARS)
        if _PROG_YEAR.match(nc or ""):
            return nc
    return ""


def build_bookings(d, orig_dir, per_file):
    """Apply corrections to copies of the originals in `orig_dir`, parse them, derive
    each booking's cohort, and return (bookings, skipped, n_context, n_applied, report).
    Nothing is silently dropped: rows with a usable group get their cohort; rows with a
    week but an unrecognised group go to an 'Unassigned' cohort and are reported."""
    os.makedirs(CORR_DIR, exist_ok=True)
    out, skipped, n_ctx, applied = [], [], 0, 0
    n_files = rows_read = 0
    detected, unmatched = {}, []
    for path in sorted(_xlsx(os.path.join(orig_dir, "*.xlsx"))):
        n_files += 1
        base = os.path.basename(path)
        per_sheet = per_file.get(base, {})
        applied += sum(len(v) for v in per_sheet.values())
        corrected = apply_to_file(path, per_sheet)
        bookings, _flags = parse_file(corrected, d)
        for b in bookings:
            if not b.week and not b.course_code:        # truly empty/template row
                continue
            rows_read += 1
            b.semester, b.academic_year = _semester_of(b.week), SPRING_AY
            coh = cohort_of(b)
            if coh is None:
                raw = (b.groups or b.groups_raw or "").strip()
                reason = "no week and no group" if not b.week else \
                    ("group not recognised" if raw else "no student group given")
                skipped.append({"file": base, "sheet": b.sheet, "code": b.course_code,
                                "groups": raw, "reason": reason, "suggest": _suggest_group(raw)})
                if not b.week:
                    continue                             # cannot place at all
                if raw:
                    unmatched.append({"sheet": b.sheet, "code": b.course_code,
                                      "raw": raw, "suggest": _suggest_group(raw)})
                b.cohort = "Unassigned"                  # still show it so the user can fix it
            else:
                b.cohort = coh
            detected[b.cohort] = detected.get(b.cohort, 0) + 1
            out.append(b)
    report = {"files": n_files, "rows_read": rows_read, "bookings_created": len(out),
              "skipped": len(skipped), "detected": detected,
              "unmatched": unmatched[:60], "skipped_list": skipped[:60]}
    return out, skipped, n_ctx, applied, report


def _write_planner(combined):
    booking_fields = [f.name for f in fields(Booking)]
    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, "bookings_2026_2027.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=booking_fields)
        w.writeheader()
        for b in combined:
            w.writerow(asdict(b))
    from .dashboard import generate as gen_dashboard
    return gen_dashboard(os.path.join(OUT, "bookings_2026_2027.csv"))


def run_upload(orig_dir, approved):
    """Import uploaded teacher files (with approved corrections) AS the planner
    source: generate the teacher/course/group lists from the files, merge them into
    the editable config, build the planner for ALL programmes found, regenerate the
    dashboard, and return a summary for the UI."""
    d = load_all()
    bk, skipped, n_ctx, applied, report = build_bookings(d, orig_dir, corrections_from_list(approved))
    from . import discover as _disc
    generated = _disc.merge_into_config(_disc.discover(bk))   # auto-generate the lists
    _dash, n = _write_planner(bk)                             # rebuild dashboard with merged config
    cohorts = sorted({b.cohort for b in bk})
    return {"ok": True, "corrections": applied, "sessions": len(bk) - n_ctx, "context": n_ctx,
            "skipped": len(skipped), "events": n, "cohorts": cohorts, "generated": generated,
            "report": report, "skipped_list": report["skipped_list"][:25]}


def run(corrections_arg=None):
    """CLI: validated spring originals + the already-cleaned autumn files."""
    d = load_all()
    per_file = load_corrections(corrections_arg)
    bk, skipped, n_ctx, applied, _report = build_bookings(d, ORIG_DIR, per_file)
    clean_bookings, _f = parse_all(d)
    autumn = [b for b in clean_bookings if b.semester == "autumn 2026"]
    combined = autumn + bk
    _dash, n = _write_planner(combined)
    print("APPLY IMPORT — validated originals → planner source")
    print(f"  corrections applied: {applied} · spring FM: {len(bk)-n_ctx} · context (KP): {n_ctx} · "
          f"autumn (cleaned): {len(autumn)} · skipped non-FM: {len(skipped)}")
    print(f"  wrote output/bookings_2026_2027.csv ({len(combined)} rows), dashboard ({n} events)")


if __name__ == "__main__":
    run(sys.argv[1] if len(sys.argv) > 1 else None)
