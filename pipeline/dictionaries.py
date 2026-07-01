"""Canonical dictionaries for courses, groups and teachers.

Per the agreed approach, raw cells are mapped onto these canonical sources:
  * courses  - config/course_master.csv (code, name, ects, notes); seeded from
               _info/.../kurskod_och_kursnamn.xlsx by scripts/seed_course_master.py
  * groups   - the 'Groups' lookup sheet in the spring request files, plus
               every valid Media-YY-X / KP-YY code generated from the known
               programme structure (gruppkoderna.txt)
  * teachers - config/teacher_aliases.csv (canonical name + nickname aliases)
               and config/teacher_typos.csv (misspelling -> canonical)
Code typos are corrected via config/course_code_fixes.csv.
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
from dataclasses import dataclass, field

import openpyxl

from .normalize import clean_text

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# When packaged with PyInstaller, read-only resources live in the bundle
# (sys._MEIPASS) while everything the app *writes* (output/, export/, _info/uploads)
# must live next to the .exe. Outside a bundle both equal the project root, so the
# normal `py`/source workflow is completely unchanged.
_FROZEN = getattr(sys, "frozen", False)
BUNDLE = getattr(sys, "_MEIPASS", ROOT) if _FROZEN else ROOT      # read-only resources
APP_DIR = os.path.dirname(sys.executable) if _FROZEN else ROOT    # writable base
CONFIG = os.path.join(BUNDLE, "config")
CONFIG_EXAMPLE = os.path.join(BUNDLE, "config.example")
# user-facing folders, always next to the app (never a Temp path)
OUTPUT_DIR = os.path.join(APP_DIR, "output")        # generated dashboard + planner CSV (internal)
EXPORT_DIR = os.path.join(APP_DIR, "export")        # final booker Excel — easy to find
IMPORT_DIR = os.path.join(APP_DIR, "import")        # drop / uploaded source files
TEMPLATES_DIR = os.path.join(APP_DIR, "templates")  # downloadable Excel templates
DATA_DIR = os.path.join(APP_DIR, "data")            # booking-request source workbooks


def cfg_path(name):
    """A config file, searched in: a config/ next to the app (lets a packaged user
    drop in their real config), then the bundled config/, then the committed generic
    config.example/ (so a public clone or fresh .exe still runs)."""
    for base in (os.path.join(APP_DIR, "config"), CONFIG, CONFIG_EXAMPLE):
        p = os.path.join(base, name)
        if os.path.exists(p):
            return p
    return os.path.join(CONFIG, name)


def data_dir():
    """Where the source booking data lives. The app reads/writes a `data/` folder
    next to itself; power users / developers can also use `_info/`. Override with the
    BA_DATA_DIR env var or config/settings.json {"data_dir": "..."}; relative paths
    resolve against the app dir. A fresh app starts with an empty data/ folder."""
    p = os.environ.get("BA_DATA_DIR")
    if not p:
        for cfg in (os.path.join(APP_DIR, "config", "settings.json"),
                    os.path.join(CONFIG, "settings.json")):
            if os.path.exists(cfg):
                try:
                    with open(cfg, encoding="utf-8-sig") as fh:
                        p = (json.load(fh) or {}).get("data_dir")
                except (ValueError, OSError):
                    p = None
                break
    if not p:
        for cand in ("_info", "data"):       # prefer one that actually holds workbooks
            full = os.path.join(APP_DIR, cand)
            if os.path.isdir(os.path.join(full, "bokningsönskemålen_2026_2027")):
                return full
        return os.path.join(APP_DIR, "data")
    return p if os.path.isabs(p) else os.path.join(APP_DIR, p)


def _settings():
    for base in (os.path.join(APP_DIR, "config"), CONFIG):
        path = os.path.join(base, "settings.json")
        if os.path.exists(path):
            try:
                with open(path, encoding="utf-8-sig") as fh:
                    return json.load(fh) or {}
            except (ValueError, OSError):
                return {}
    return {}


INFO = data_dir()
COURSE_MASTER = cfg_path("course_master.csv")
COURSE_FIXES = cfg_path("course_code_fixes.csv")
TEACHER_FILE = cfg_path("teacher_aliases.csv")
TEACHER_TYPOS = cfg_path("teacher_typos.csv")
GROUPS_SOURCE = os.path.join(INFO, "bokningsönskemålen_2026_2027", "våren_2027", "media-25-VT-2027.xlsx")

# Arcada programmes (code -> name). The reference list, used as defaults and for the
# "add program" dropdown. Editable in-app (stored in settings.json "programs").
DEFAULT_PROGRAMS = {
    "Media": "Film och media", "KP": "Kulturproducent", "FT": "Fysioterapi",
    "SJ": "Sjukskötare", "HV": "Hälsovårdare", "BM": "Barnmorska", "FV": "Förstavård",
    "ET": "Ergoterapi", "IT": "Informationsteknik", "IB": "International Business",
    "FE": "Företagsekonomi",
}
DEFAULT_MEDIA_TRACKS = {"F": "Foto", "L": "Ljud", "M": "Manus", "O": "Online", "P": "Producing"}


def active_cohort_years():
    """The current intake years to plan — newest `window` cohorts (default 4). Base
    defaults to this calendar year (autumn 2026 -> 26, giving 26/25/24/23) and rolls
    forward automatically each year; both base and window are overridable in settings."""
    import datetime
    s = _settings()
    base = int(s.get("active_base_year", datetime.datetime.now().year % 100))
    window = max(1, int(s.get("active_window", 4)))
    return [f"{(base - i) % 100:02d}" for i in range(window)]


def current_academic_year():
    """e.g. '2026-2027' from the newest active intake year."""
    yrs = active_cohort_years()
    base = int(yrs[0]) if yrs else datetime.datetime.now().year % 100
    return f"20{base:02d}-20{base + 1:02d}"


def programs():
    s = _settings()
    return dict(s.get("programs") or DEFAULT_PROGRAMS)


def media_tracks():
    return dict(_settings().get("specializations") or DEFAULT_MEDIA_TRACKS)


_S = _settings()
SPEC_LETTERS = "".join(media_tracks().keys()) or "FLMOP"
COHORT_YEARS = active_cohort_years()                  # newest active intake years


@dataclass
class Dictionaries:
    courses: dict = field(default_factory=dict)          # code -> name
    course_ects: dict = field(default_factory=dict)      # code -> ects (str, may be "")
    course_notes: dict = field(default_factory=dict)     # code -> notes
    code_fixes: dict = field(default_factory=dict)       # wrong -> (correct, reason)
    groups: dict = field(default_factory=dict)           # code -> description
    teacher_alias: dict = field(default_factory=dict)    # lower(alias) -> canonical
    teachers: list = field(default_factory=list)         # canonical names
    _name_re: object = None                              # regex over known names/aliases

    # ---- course lookups -------------------------------------------------- #
    def lookup_course(self, code: str):
        """Return (canonical_code, name, note).

        Applies code_fixes, then tries an exact match, then a
        zero-pad-insensitive match (so 'TV-2-34' resolves to 'TV-2-034')."""
        note = ""
        if code in self.code_fixes:
            correct, reason = self.code_fixes[code]
            note = f"code corrected {code} -> {correct} ({reason})"
            code = correct
        if code in self.courses:
            return code, self.courses[code], note
        key = _padkey(code)
        if key:
            for c, name in self.courses.items():
                if _padkey(c) == key:
                    pad = f"code '{code}' matched '{c}' ignoring zero-padding"
                    return c, name, (note + "; " + pad if note else pad)
        miss = f"course code '{code}' not in dictionary"
        return None, None, (note + "; " + miss if note else miss)

    def ects(self, code):
        return self.course_ects.get(code, "")

    # ---- teacher lookups ------------------------------------------------- #
    def lookup_teacher(self, name: str):
        """Return (canonical_name, note). Unknown names are kept as-is and,
        per the agreed rule, NOT flagged (note is '')."""
        cleaned = clean_text(name)
        canonical = self.teacher_alias.get(cleaned.lower())
        if canonical:
            note = "" if canonical == cleaned else f"'{cleaned}' -> '{canonical}'"
            return canonical, note
        return cleaned, ""  # unknown -> ignore

    def split_known(self, token: str):
        """If an unrecognized token actually contains several known names run
        together (e.g. 'Alex Smith Robin Lee'), return the canonical names;
        otherwise None."""
        if not self._name_re:
            return None
        found = [self.teacher_alias[m.group(0).lower()] for m in self._name_re.finditer(token)]
        # de-dup while preserving order
        seen, out = set(), []
        for n in found:
            if n not in seen:
                seen.add(n); out.append(n)
        return out if len(out) >= 2 else None

    # ---- group lookups --------------------------------------------------- #
    def is_known_group(self, code: str) -> bool:
        return code in self.groups


def _read_csv(path):
    """Read a config CSV tolerantly: try UTF-8 (with/without BOM), fall back to
    Windows-1252, since hand-edited files may be saved in either encoding."""
    for enc in ("utf-8-sig", "cp1252"):
        try:
            with open(path, encoding=enc) as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError:
            continue
    with open(path, encoding="utf-8", errors="replace") as f:
        return list(csv.DictReader(f))


def _padkey(code: str):
    parts = code.split("-")
    if len(parts) < 2 or not parts[-1].isdigit():
        return None
    return ("-".join(parts[:-1]).upper(), int(parts[-1]))


def load_courses():
    courses, ects, notes = {}, {}, {}
    for row in _read_csv(cfg_path("course_master.csv")):   # resolve at call time, not import time
        code = clean_text(row["code"]).upper()
        if not code:
            continue
        courses[code] = clean_text(row.get("name"))
        ects[code] = clean_text(row.get("ects"))
        notes[code] = clean_text(row.get("notes"))
    return courses, ects, notes


def load_code_fixes():
    fixes = {}
    path = cfg_path("course_code_fixes.csv")
    if not os.path.exists(path):
        return fixes
    for row in _read_csv(path):
        wrong = clean_text(row["wrong_code"]).upper()
        correct = clean_text(row["correct_code"]).upper()
        if wrong and correct:
            fixes[wrong] = (correct, clean_text(row.get("reason")))
    return fixes


def generate_group_codes():
    """All group codes for the active programmes × active intake years (CODE-YY),
    plus Media tracks (Media-YY-F/L/…) and any manual/imported extra groups.
    Returns {code: programme name}."""
    s = _settings()
    progs = programs()
    active = set(s.get("active_programs") or progs.keys())
    tracks = media_tracks()
    years = active_cohort_years()
    out = {}
    for code, name in progs.items():
        if code not in active:
            continue
        for yy in years:
            out[f"{code}-{yy}"] = name
            if code == "Media":
                for t, tname in tracks.items():
                    out[f"Media-{yy}-{t}"] = f"{name} ({tname})"
    for g in (s.get("extra_groups") or []):
        g = (g or "").strip()
        if g:
            out.setdefault(g, "(added manually)")
    return out


def load_groups():
    groups = generate_group_codes()
    try:
        wb = openpyxl.load_workbook(GROUPS_SOURCE, data_only=True, read_only=True)
        if "Groups" in wb.sheetnames:
            for row in wb["Groups"].iter_rows(values_only=True):
                if row and row[0]:
                    code = clean_text(row[0])
                    desc = clean_text(row[1]) if len(row) > 1 and row[1] else ""
                    groups.setdefault(code, desc or "(cross-programme)")
        wb.close()
    except FileNotFoundError:
        pass
    return groups


def load_teachers():
    alias, canon = {}, []
    for row in _read_csv(cfg_path("teacher_aliases.csv")):   # resolve at call time
        name = (row.get("canonical_name") or "").strip()
        if not name:
            continue
        canon.append(name)
        alias[name.lower()] = name
        for a in (row.get("aliases") or "").split(";"):
            a = a.strip()
            if a:
                alias[a.lower()] = name
    typos = cfg_path("teacher_typos.csv")
    if os.path.exists(typos):
        for row in _read_csv(typos):
            wrong = (row.get("wrong") or "").strip()
            correct = (row.get("correct") or "").strip()
            if wrong and correct:
                alias[wrong.lower()] = correct
    return alias, canon


def load_all() -> Dictionaries:
    courses, ects, notes = load_courses()
    alias, canon = load_teachers()
    # Regex of all known names+aliases, longest first, for concatenation splits.
    keys = sorted(alias.keys(), key=len, reverse=True)
    name_re = re.compile("|".join(re.escape(k) for k in keys), re.I) if keys else None
    return Dictionaries(
        courses=courses, course_ects=ects, course_notes=notes,
        code_fixes=load_code_fixes(), groups=load_groups(),
        teacher_alias=alias, teachers=canon, _name_re=name_re,
    )
