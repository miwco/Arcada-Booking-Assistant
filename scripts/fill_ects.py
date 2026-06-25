"""Derive course ECTS from the curriculum workbook and write them into
config/course_master.csv.

The curriculum (utbildningsupplägget_FM_2026-2027.xlsx) draws each course as a
merged box. Box size encodes ECTS: a box of N grid cells (rowspan x colspan)
maps to ECTS = round(N/2)*5, i.e. 2 cells -> 5, 4 -> 10, 6 -> 15. (A "normal"
box is 2x1 and a "long" box is 1x2; both are 2 cells = 5 ECTS.)

Courses that are not mandatory are not drawn in the curriculum, so a few ECTS
values are supplied by hand in OVERRIDES.
"""
import csv
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
from pipeline.normalize import clean_text  # noqa: E402

CURRICULUM = os.path.join(ROOT, "_info", "om_kurserna_grupperna_utbildningen",
                          "utbildningsupplägget_FM_2026-2027.xlsx")
MASTER = os.path.join(ROOT, "config", "course_master.csv")
CODE_RE = re.compile(r"[A-ZÅÄÖ]{2,3}-\d-\d{2,3}")

# Codes used in requests but corrected elsewhere; their ECTS follows the target.
CODE_FIXES = {"MK-2-172": "MK-2-131", "ML-2-125": "MK-2-125"}
# Non-mandatory courses not drawn in the curriculum (ECTS given by the user).
OVERRIDES = {"MK-2-155": "15", "FM-3-007": "5", "YH-4-032": "5"}
# Best-guess ECTS for courses not drawn in the curriculum (verify these).
GUESSES = {
    "SP-SV-001": "5", "SP-FI-001": "5", "TV-2-002": "5",
    "FM-3-005": "5", "FM-3-004": "5", "FM-2-014": "15", "MK-2-115": "5",
}


def cells_to_ects(cells):
    return max(5, min(15, round(cells / 2) * 5))


def curriculum_ects():
    wb = openpyxl.load_workbook(CURRICULUM, data_only=True)
    ects, names = {}, {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for mr in ws.merged_cells.ranges:
            v = ws.cell(mr.min_row, mr.min_col).value
            if isinstance(v, str) and CODE_RE.search(v):
                code = CODE_RE.search(v).group(0)
                cells = (mr.max_row - mr.min_row + 1) * (mr.max_col - mr.min_col + 1)
                ects.setdefault(code, cells_to_ects(cells))
                # name = text after the code on the same cell
                nm = clean_text(v.split(code, 1)[-1]).lstrip(",: ").strip()
                if nm:
                    names.setdefault(code, nm)
    wb.close()
    return ects, names


def main():
    cur_ects, cur_names = curriculum_ects()

    def ects_for(code):
        if code in OVERRIDES:
            return OVERRIDES[code]
        if code in cur_ects:
            return str(cur_ects[code])
        return ""

    # Load existing master.
    rows, seen = [], set()
    with open(MASTER, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            code = r["code"].strip().upper()
            seen.add(code)
            if not r.get("ects"):
                r["ects"] = ects_for(code)
            if not r.get("ects") and code in GUESSES:
                r["ects"] = GUESSES[code]
                note = (r.get("notes") or "").strip()
                r["notes"] = (note + "; " if note else "") + "ECTS guessed - verify"
            rows.append(r)

    # Add curriculum courses missing from the master (e.g. MK-2-173, MK-2-179).
    for code in sorted(cur_names):
        if code in seen or code in CODE_FIXES:  # skip typo-codes
            continue
        rows.append({"code": code, "name": cur_names[code],
                     "ects": ects_for(code), "notes": "added from curriculum"})

    rows.sort(key=lambda r: r["code"])
    with open(MASTER, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["code", "name", "ects", "notes"])
        w.writeheader()
        w.writerows(rows)

    filled = sum(1 for r in rows if r["ects"])
    print(f"course_master.csv: {len(rows)} courses, {filled} with ECTS, "
          f"{len(rows)-filled} still blank")
    blanks = [r["code"] for r in rows if not r["ects"]]
    if blanks:
        print("  still blank:", ", ".join(blanks))


if __name__ == "__main__":
    main()
